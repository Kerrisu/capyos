"""
CapyOS Backend - logica_escala.py (adaptado da versão desktop)

Principais mudanças em relação ao original:
1. Removida a dependência de caminho de .exe (sys._MEIPASS) para achar credenciais.
2. Credenciais agora podem vir de:
   a) Variável de ambiente GOOGLE_CREDENTIALS_JSON (conteúdo do credenciais.json como string) -> USO EM PRODUÇÃO
   b) Arquivo local "credenciais.json" na raiz do projeto -> USO EM DESENVOLVIMENTO LOCAL (nunca commitar!)
3. callback_progresso continua existindo (mantém compatibilidade com a lógica original),
   mas agora ele só imprime debug — quem realmente vai acompanhar progresso pela rota
   HTTP receberá o resultado final de uma vez (streaming de progresso fica pra uma etapa futura, se quiser).
"""

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
import json
import re
import unicodedata
import requests
import openpyxl
from io import BytesIO

# --- CONFIGURAÇÕES ---
HORARIOS_PADRAO = ["13:15H", "14:00H", "14:45H", "15:30H", "16:15H", "17:00H", "17:45H"]

DEBUG_TAG = "🔧[CAPYOS-DEBUG]"

DRIVE_API_BASE = "https://www.googleapis.com/drive/v3"
MIMETYPE_SHEETS_NATIVO = "application/vnd.google-apps.spreadsheet"


def carregar_credenciais_dict():
    """
    Retorna o dicionário de credenciais do Google, buscando primeiro na
    variável de ambiente (produção) e depois no arquivo local (desenvolvimento).
    """
    creds_env = os.environ.get("GOOGLE_CREDENTIALS_JSON")

    if creds_env:
        print(f"{DEBUG_TAG} Credenciais encontradas via variável de ambiente GOOGLE_CREDENTIALS_JSON.")
        try:
            return json.loads(creds_env)
        except json.JSONDecodeError as e:
            print(f"{DEBUG_TAG} ERRO: GOOGLE_CREDENTIALS_JSON não é um JSON válido: {e}")
            raise

    # Fallback para desenvolvimento local
    caminho_local = os.path.join(os.path.dirname(os.path.abspath(__file__)), "credenciais.json")
    print(f"{DEBUG_TAG} Variável de ambiente não encontrada. Tentando arquivo local: {caminho_local}")

    if os.path.exists(caminho_local):
        print(f"{DEBUG_TAG} Arquivo local de credenciais encontrado, carregando...")
        with open(caminho_local, "r", encoding="utf-8") as f:
            return json.load(f)

    print(f"{DEBUG_TAG} ERRO: Nenhuma credencial encontrada (nem env var, nem arquivo local).")
    raise FileNotFoundError(
        "Credenciais do Google não encontradas. Configure GOOGLE_CREDENTIALS_JSON "
        "ou coloque um credenciais.json na raiz do projeto (apenas em dev local)."
    )


def conectar_google_sheets(url_planilha):
    print(f"{DEBUG_TAG} Iniciando conexão com Google Sheets para URL: {url_planilha}")
    escopos = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

    creds_dict = carregar_credenciais_dict()
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, escopos)
    print(f"{DEBUG_TAG} Credenciais carregadas com sucesso. Autorizando cliente gspread...")

    client = gspread.authorize(creds)
    print(f"{DEBUG_TAG} Cliente gspread autorizado. Abrindo planilha...")

    planilha = client.open_by_url(url_planilha)
    print(f"{DEBUG_TAG} Planilha aberta com sucesso: '{planilha.title}'")
    return planilha


def _extrair_id_arquivo(url_ou_id):
    """Extrai o ID do arquivo a partir de uma URL do Google Drive/Sheets.
    Se já vier só o ID (sem barras), devolve como está."""
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", url_ou_id)
    return match.group(1) if match else url_ou_id


def _obter_headers_autenticados():
    """
    Gera um header Authorization Bearer a partir das MESMAS credenciais de
    serviço já usadas pelo gspread, pra chamadas cruas na Drive API (que o
    gspread não cobre): checar o mimeType do arquivo e, se for um .xlsx
    real, baixar os bytes brutos.

    Não precisa de nenhum escopo novo — "spreadsheets" + "drive" (já
    usados em conectar_google_sheets) cobrem leitura de metadados e
    download de conteúdo.
    """
    escopos = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds_dict = carregar_credenciais_dict()
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, escopos)
    token = creds.get_access_token().access_token
    return {"Authorization": f"Bearer {token}"}


def _detectar_mimetype_arquivo(file_id, headers):
    resp = requests.get(
        f"{DRIVE_API_BASE}/files/{file_id}",
        headers=headers,
        params={"fields": "mimeType"},
    )
    resp.raise_for_status()
    return resp.json().get("mimeType", "")


def _hex_para_rgb_float(hex_cor):
    """
    Converte uma cor hex do openpyxl (formatos 'RRGGBB' ou 'AARRGGBB', o
    canal alfa quando presente é ignorado) pra uma tupla (r, g, b) de
    floats 0.0-1.0 — o MESMO formato que o Google Sheets API usa em
    effectiveFormat.backgroundColor. Isso é o que permite reaproveitar
    _cor_bate() sem nenhuma mudança pros dois tipos de fonte.
    """
    if not hex_cor or len(hex_cor) < 6:
        return (1.0, 1.0, 1.0)
    hex_rgb = hex_cor[-6:]
    try:
        r = int(hex_rgb[0:2], 16) / 255.0
        g = int(hex_rgb[2:4], 16) / 255.0
        b = int(hex_rgb[4:6], 16) / 255.0
        return (r, g, b)
    except ValueError:
        return (1.0, 1.0, 1.0)


def _construir_rows_data_de_xlsx(conteudo_bytes, nome_aba):
    """
    Lê um .xlsx cru (bytes, baixado direto da Drive API) com openpyxl e
    devolve os dados no MESMO formato que fetch_sheet_metadata() do Google
    Sheets API devolveria: uma lista de linhas, cada uma com 'values' ->
    lista de células com 'formattedValue' e
    'effectiveFormat.backgroundColor'. Isso é o que permite que
    processar_escala()/inspecionar_cores() continuem exatamente iguais,
    sem precisar de nenhum "if fonte == xlsx" espalhado pelo parsing.

    data_only=True: pega o VALOR calculado das células (não a fórmula
    crua), igual ao que formattedValue do Sheets API já devolvia.
    """
    workbook = openpyxl.load_workbook(BytesIO(conteudo_bytes), data_only=True)

    if nome_aba in workbook.sheetnames:
        aba = workbook[nome_aba]
    else:
        aba = workbook.worksheets[0]
        print(f"{DEBUG_TAG} Aba '{nome_aba}' não encontrada no .xlsx, usando a primeira disponível: '{aba.title}'")

    rows_data = []
    for row in aba.iter_rows():
        values = []
        for cell in row:
            valor = "" if cell.value is None else str(cell.value)

            r, g, b = (1.0, 1.0, 1.0)
            fill = cell.fill
            if fill is not None and fill.patternType == "solid":
                cor = fill.fgColor
                if cor is not None and cor.type == "rgb" and cor.rgb and cor.rgb not in ("00000000", None):
                    r, g, b = _hex_para_rgb_float(cor.rgb)

            values.append({
                "formattedValue": valor,
                "effectiveFormat": {"backgroundColor": {"red": r, "green": g, "blue": b}},
            })
        rows_data.append({"values": values})

    return rows_data, aba.title


def _obter_rows_data(url_planilha, nome_aba):
    """
    Ponto único de leitura da planilha de origem. Detecta automaticamente
    se o arquivo é uma planilha NATIVA (Google Sheets — como a planilha de
    teste usada hoje) ou um .xlsx REAL (como o Direcionamento), e lê pelo
    caminho certo em cada caso:

    - Nativa  -> gspread + Sheets API (fetch_sheet_metadata), igual sempre foi.
    - .xlsx   -> download bruto via Drive API (files.get?alt=media, que
                 funciona com acesso de leitor/link — não cria nem copia
                 nada, então não esbarra na cota-zero de Drive que
                 service accounts têm) + parsing local com openpyxl.

    Em ambos os casos devolve (rows_data, titulo_aba_usada) no MESMO
    formato, pra processar_escala/inspecionar_cores não precisarem saber
    qual dos dois caminhos foi usado.
    """
    headers = _obter_headers_autenticados()
    file_id = _extrair_id_arquivo(url_planilha)
    mimetype = _detectar_mimetype_arquivo(file_id, headers)

    if mimetype == MIMETYPE_SHEETS_NATIVO:
        print(f"{DEBUG_TAG} Arquivo é planilha nativa (mimeType={mimetype}). Lendo via Sheets API.")
        planilha = conectar_google_sheets(url_planilha)
        try:
            aba = planilha.worksheet(nome_aba)
        except Exception:
            aba = planilha.get_worksheet(0)
            print(f"{DEBUG_TAG} Aba '{nome_aba}' não encontrada, usando a primeira disponível: '{aba.title}'")

        fmt = aba.spreadsheet.fetch_sheet_metadata({
            "ranges": [aba.title],
            "includeGridData": True,
            "fields": "sheets.properties.title,sheets.data.rowData.values.formattedValue,sheets.data.rowData.values.effectiveFormat.backgroundColor",
        })
        sheet_data = [s for s in fmt['sheets'] if s['properties']['title'] == aba.title][0]
        rows_data = sheet_data['data'][0].get('rowData', [])
        return rows_data, aba.title

    else:
        print(f"{DEBUG_TAG} Arquivo é .xlsx real (mimeType={mimetype}). Baixando bytes brutos...")
        resp = requests.get(
            f"{DRIVE_API_BASE}/files/{file_id}",
            headers=headers,
            params={"alt": "media"},
        )
        resp.raise_for_status()
        return _construir_rows_data_de_xlsx(resp.content, nome_aba)


def _listar_nomes_abas(url_planilha):
    """Lista os nomes das abas da planilha de origem, nativa ou .xlsx real."""
    headers = _obter_headers_autenticados()
    file_id = _extrair_id_arquivo(url_planilha)
    mimetype = _detectar_mimetype_arquivo(file_id, headers)

    if mimetype == MIMETYPE_SHEETS_NATIVO:
        planilha = conectar_google_sheets(url_planilha)
        return [aba.title for aba in planilha.worksheets()]
    else:
        resp = requests.get(
            f"{DRIVE_API_BASE}/files/{file_id}",
            headers=headers,
            params={"alt": "media"},
        )
        resp.raise_for_status()
        workbook = openpyxl.load_workbook(BytesIO(resp.content), read_only=True)
        return workbook.sheetnames


# FUNÇÃO PARA LISTAR ABAS DISPONÍVEIS NA PLANILHA
def listar_abas(url_planilha):
    """Retorna uma lista com os nomes de todas as abas da planilha"""
    print(f"{DEBUG_TAG} listar_abas() chamada com url={url_planilha}")
    try:
        abas = _listar_nomes_abas(url_planilha)
        print(f"{DEBUG_TAG} Abas encontradas: {abas}")
        return abas
    except Exception as e:
        print(f"{DEBUG_TAG} ERRO ao listar abas: {e}")
        return []


# --- PROCESSAMENTO DA ESCALA ---

# --- CORES DE REFERÊNCIA (calibradas em 06/08/2026 com a planilha real) ---
# Usamos combinação EXATA (com pequena tolerância) em vez de faixas soltas,
# porque a planilha ganhou mais cores (rosa/magenta, azul, cinza) e faixas
# soltas tipo "g > 0.65" acabavam capturando cores erradas por engano
# (ex: o azul da PISCINA caía dentro da faixa de verde).
COR_VERDE = (0.4, 1.0, 0.6)      # REFERÊNCIA
COR_AMARELO = (1.0, 1.0, 0.4)    # SUPRIDA
# Rosa claro: usado temporariamente pro time marcar REFERÊNCIA de assistidos
# TRANSFERIDOS (confirmado com o Ken em 23/08). É a mesma categoria clínica
# do verde, só com cor diferente por causa do processo de transferência —
# a expectativa é que vire verde puro nos próximos dias, mas tratamos como
# equivalente desde já pra não perder esses assistidos na alocação.
COR_REFERENCIA_TRANSFERIDO = (1.0, 0.78, 0.808)
TOLERANCIA_COR = 0.05


def _cor_bate(r, g, b, alvo):
    """Confere se a cor (r,g,b) está bem próxima da cor alvo, com tolerância pequena."""
    return (
        abs(r - alvo[0]) <= TOLERANCIA_COR and
        abs(g - alvo[1]) <= TOLERANCIA_COR and
        abs(b - alvo[2]) <= TOLERANCIA_COR
    )


def _normalizar(texto):
    """Maiúsculo e sem acento, pra comparações de texto não falharem por causa de á/ã/ç etc."""
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).upper().strip()


def _normalizar_horario(texto):
    """
    Normaliza texto de horário pra comparação: maiúsculo, sem 'H' e sem
    NENHUM espaço (início, fim ou meio).

    IMPORTANTE (lição da Parte 2, item 4): existia uma bug de "17:00H"
    sumindo silenciosamente nas sextas-feiras porque a planilha tinha
    "17:00 H" (espaço extra antes do H) em vez de "17:00H". A comparação
    antiga usava duas implementações de normalização diferentes — uma
    com .strip() no fim (pra montar a lista de referência) e outra sem
    (pra ler cada linha da planilha) — que podiam divergir exatamente
    nesse tipo de caso. Agora existe UMA função só, usada nos dois
    lugares, e ela remove QUALQUER espaço (não só nas pontas), pra
    erros de digitação assim não quebrarem mais a leitura da escala.
    """
    if not texto:
        return ""
    return "".join(texto.upper().replace("H", "").split())


def _remover_sufixo_parenteses(texto):
    """
    Remove qualquer sufixo entre parênteses do nome de um assistido
    (ex: "FELIPE SALEMI (SUPRIR AT)" -> "FELIPE SALEMI", "MELINA VITORIA
    (SUPRIR DAYSI PP)" -> "MELINA VITORIA").

    IMPORTANTE (bug encontrado na Parte 2): sem isso, o nome com
    parêntese não batia com nada — nem com o cadastro em `pacientes_db`,
    nem com o dicionário `apelidos` em `distribuir_salas_ia` — porque
    todos eles comparam a string exata do nome. Resultado: o assistido
    perdia silenciosamente sala fixa, prioridade clínica e grupo de
    compatibilidade (`grupo_match`), sem nenhum erro visível.
    Aplicado aqui, o mais cedo possível na leitura da célula, pra ser a
    única fonte de verdade — tudo que vem depois (banco, apelidos,
    super_grupo, texto final) já trabalha só com o nome limpo.

    Por enquanto o conteúdo dentro do parêntese (ex: "SUPRIR AT") não é
    guardado em lugar nenhum — só é descartado.
    """
    if not texto:
        return texto
    sem_parenteses = re.sub(r"\([^)]*\)", "", texto)
    return " ".join(sem_parenteses.split())


def processar_escala(url_planilha, callback_progresso, nome_aba):
    """
    callback_progresso: função que recebe (valor_float, texto_status).
    Na versão desktop, ela mexia direto na barra do Tkinter.
    Aqui, por padrão, ela só imprime debug — mas mantém a assinatura
    para não quebrar compatibilidade e permitir um futuro sistema de
    progresso em tempo real via WebSocket, se você quiser.
    """
    try:
        callback_progresso(0.1, "Conectando e baixando metadados...")
        print(f"{DEBUG_TAG} processar_escala() iniciada. nome_aba={nome_aba}")

        rows_data, titulo_aba_usada = _obter_rows_data(url_planilha, nome_aba)
        print(f"{DEBUG_TAG} Aba usada: '{titulo_aba_usada}'. {len(rows_data)} linhas baixadas da planilha.")

        pacientes_encontrados = []
        registrados = set()
        horarios_limpos = [_normalizar_horario(h) for h in HORARIOS_PADRAO]

        # Mapa coluna -> nome do profissional daquele bloco. É atualizado toda
        # vez que encontramos uma linha "HORÁRIO" (a linha ANTERIOR a ela tem
        # os nomes dos profissionais). Isso é necessário pra regra do Jorge
        # (assistidos dele vão sempre pra ABA 04), já que a ordem das colunas
        # muda de bloco pra bloco no dia.
        mapa_coluna_profissional = {}
        linha_anterior_values = None

        total_linhas = len(rows_data)
        callback_progresso(0.3, "Analisando cores dos pacientes...")

        for i, row in enumerate(rows_data):
            if i % 5 == 0:
                prog = 0.3 + (float(i) / float(total_linhas)) * 0.6 if total_linhas else 0.3
                callback_progresso(prog, f"Processando linha {i}...")

            values = row.get('values', [])
            if not values:
                linha_anterior_values = values
                continue

            texto_coluna_a_bruto = values[0].get('formattedValue', '').strip()
            texto_normalizado = _normalizar(texto_coluna_a_bruto)

            if texto_normalizado == "HORARIO":
                novo_mapa = {}
                if linha_anterior_values:
                    for j, cell in enumerate(linha_anterior_values):
                        if j == 0:
                            continue
                        nome_prof = cell.get('formattedValue', '').strip()
                        if nome_prof:
                            novo_mapa[j] = nome_prof
                if novo_mapa:
                    mapa_coluna_profissional = novo_mapa
                    print(f"{DEBUG_TAG} Novo bloco de profissionais detectado: {mapa_coluna_profissional}")
                linha_anterior_values = values
                continue

            texto_coluna_a = _normalizar_horario(texto_coluna_a_bruto)

            if texto_coluna_a not in horarios_limpos:
                linha_anterior_values = values
                continue

            for j, cell in enumerate(values):
                if j == 0:
                    continue

                valor = cell.get('formattedValue', '').strip()
                if not valor:
                    continue

                # Remove sufixo entre parênteses (ex: "(SUPRIR AT)") ANTES de
                # qualquer normalização ou registro — precisa ser a primeira
                # coisa feita com o texto bruto da célula. Ver docstring de
                # _remover_sufixo_parenteses() pro contexto do bug.
                valor = _remover_sufixo_parenteses(valor)
                if not valor:
                    continue

                nome_limpo = " ".join(valor.upper().split())
                chave_unica = (nome_limpo, texto_coluna_a)

                if chave_unica in registrados:
                    continue

                cor_data = cell.get('effectiveFormat', {}).get('backgroundColor', {})
                r = cor_data.get('red', 0)
                g = cor_data.get('green', 0)
                b = cor_data.get('blue', 0)

                is_verde = _cor_bate(r, g, b, COR_VERDE) or _cor_bate(r, g, b, COR_REFERENCIA_TRANSFERIDO)
                is_amarelo = _cor_bate(r, g, b, COR_AMARELO)

                if is_verde or is_amarelo:
                    pacientes_encontrados.append({
                        "nome": nome_limpo,
                        "horario": texto_coluna_a + "H",
                        "tipo": "REFERENCIA" if is_verde else "SUPRIDA",
                        "profissional": mapa_coluna_profissional.get(j)
                    })
                    registrados.add(chave_unica)

            linha_anterior_values = values

        callback_progresso(0.9, "Organizando lista final...")
        print(f"{DEBUG_TAG} processar_escala() concluída. {len(pacientes_encontrados)} pacientes encontrados.")
        return pacientes_encontrados

    except Exception as e:
        print(f"{DEBUG_TAG} ERRO em processar_escala(): {e}")
        return f"Erro no processamento: {str(e)}"


# --- DIAGNÓSTICO DE CORES (usado quando a planilha muda o padrão visual) ---

def inspecionar_cores(url_planilha, nome_aba):
    """
    Não classifica nada. Só varre as células preenchidas nos horários padrão
    e devolve a cor RGB crua de cada uma, pra gente conseguir recalibrar
    is_verde/is_amarelo quando a planilha mudar de tom.
    """
    print(f"{DEBUG_TAG} inspecionar_cores() iniciada. nome_aba={nome_aba}")
    try:
        rows_data, titulo_aba_usada = _obter_rows_data(url_planilha, nome_aba)
        print(f"{DEBUG_TAG} Aba usada: '{titulo_aba_usada}'.")

        horarios_limpos = [_normalizar_horario(h) for h in HORARIOS_PADRAO]
        amostras = []

        for row in rows_data:
            values = row.get('values', [])
            if not values:
                continue

            texto_coluna_a = _normalizar_horario(values[0].get('formattedValue', ''))
            if texto_coluna_a not in horarios_limpos:
                continue

            for j, cell in enumerate(values):
                if j == 0:
                    continue
                valor = cell.get('formattedValue', '').strip()
                if not valor:
                    continue

                # Mesmo tratamento de processar_escala, pra o diagnóstico de
                # cores refletir o nome real usado no processamento (uma
                # fonte de verdade só).
                valor = _remover_sufixo_parenteses(valor)
                if not valor:
                    continue

                cor_data = cell.get('effectiveFormat', {}).get('backgroundColor', {})
                r = round(cor_data.get('red', 0), 3)
                g = round(cor_data.get('green', 0), 3)
                b = round(cor_data.get('blue', 0), 3)

                veredito = "VERDE" if (_cor_bate(r, g, b, COR_VERDE) or _cor_bate(r, g, b, COR_REFERENCIA_TRANSFERIDO)) else (
                    "AMARELO" if _cor_bate(r, g, b, COR_AMARELO) else "ignorado"
                )

                amostras.append({
                    "nome": " ".join(valor.upper().split()),
                    "horario": texto_coluna_a + "H",
                    "r": r,
                    "g": g,
                    "b": b,
                    "veredito": veredito
                })

        print(f"{DEBUG_TAG} inspecionar_cores() concluída. {len(amostras)} células amostradas.")
        return amostras

    except Exception as e:
        print(f"{DEBUG_TAG} ERRO em inspecionar_cores(): {e}")
        return f"Erro: {str(e)}"


# --- FUNÇÃO DE DISTRIBUIÇÃO DE SALAS (IA) ---

def distribuir_salas_ia(lista_pacientes, configuracoes):
    print(f"{DEBUG_TAG} distribuir_salas_ia() iniciada com {len(lista_pacientes) if isinstance(lista_pacientes, list) else 'N/A'} pacientes.")

    pacientes_db = configuracoes.get("pacientes", {})
    regras_gerais = configuracoes.get("configuracoes_gerais", {})

    # REMOVIDO (Parte 2, item "grupo_match"): a lista `super_grupo` que
    # existia aqui hardcoded foi levantada manualmente há um tempo e ficou
    # desatualizada (faltavam vários assistidos que dividem sala tranquilo,
    # e sobrava gente que na real não deveria estar lá). Ela foi
    # substituída pelo campo `grupo_match` no banco (Neon), que é
    # atualizável pela tela de Gerenciar Assistidos sem precisar mexer em
    # código. Todo assistido com `grupo_match` preenchido com o mesmo valor
    # (hoje: "flexivel") é tratado como parte do grupo que divide sala com
    # qualquer perfil compatível — ver uso mais abaixo, na hora de decidir
    # se alguém pode entrar numa sala já ocupada.

    # ATUALIZADO (Parte 2, item 3 - padronização de nomes entre planilha e
    # banco). Cada chave é uma variante encontrada na planilha real
    # (Agenda dos Aplicadores / DIRECIONAMENTO) que difere do nome
    # cadastrado no banco (Neon) — seja por letra faltando/trocada, seja
    # por diferença de acentuação. O valor é SEMPRE a grafia exata que
    # está (ou deveria estar) cadastrada em `pacientes_db`.
    #
    # ⚠️ ATENÇÃO — dois nomes abaixo dependem de correção manual no banco:
    # o registro de "JOÃO WILLAMS" precisa ser renomeado pra "JOÃO WILLIAMS"
    # e o de "WILLIAN ALVES" precisa ser renomeado pra "WILLIAM ALVES" na
    # tela Gerenciar Assistidos. Até isso ser feito, esses dois apelidos
    # resolvem o nome certo mas não encontram a config no banco (sala
    # fixa, divide_sala etc. ficam vazios pra eles).
    apelidos = {
        "MALU": "MARIA LUIZA",

        "LAVINIA F": "LAVINIA FIGUEIRAS",
        "LAVINIA FIGUEIRA": "LAVINIA FIGUEIRAS",
        "LAVÍNIA FIGUEIRAS": "LAVINIA FIGUEIRAS",
        "LAVÍNIA FIGUEIRA": "LAVINIA FIGUEIRAS",

        "LUCAS E": "LUCAS EMMANUEL",
        "LUCAS EMANUEL": "LUCAS EMMANUEL",

        "ANTONY V": "ANTHONY VINICIUS",
        "ANTHONY VINCIUS": "ANTHONY VINICIUS",

        "PEDRO ACIOLY": "PEDRO ACCIOLY",

        # WILLIAM ALVES / JOÃO WILLIAMS: ver aviso acima sobre correção
        # pendente no banco.
        "WILLIAM": "WILLIAM ALVES",
        "WILLIAN": "WILLIAM ALVES",
        "WILLIAN ALVES": "WILLIAM ALVES",

        "JOÃO WILLAMS": "JOÃO WILLIAMS",
        "JOÃO WILLIANS": "JOÃO WILLIAMS",

        "JOÃO MATEUS": "JOÃO MATHEUS",

        "LUCCA CALVACANTI": "LUCCA CAVALCANTI",

        "THEO VITOR": "THEO VICTOR",

        "BEATRIZ ARAÚJO": "BEATRIZ ARAUJO",
        "BEATRIZ ARÚJO": "BEATRIZ ARAUJO",

        "LAURA ARAUJO": "LAURA ARAÚJO",

        "LÍLIA MELO": "LILIA MELO",

        "MARIA JULIA": "MARIA JÚLIA",

        "MARIA VITORIA": "MARIA VITÓRIA",

        "MELINA VITORIA": "MELINA VITÓRIA",
    }

    salas_terreo = regras_gerais.get("ordem_salas_terreo", [])
    bloqueadas = regras_gerais.get("salas_bloqueadas", [])

    # REMOVIDO ("no flow" - item 2): a lista fixa `ordem_salas_preferencial`
    # que existia aqui (ex: sempre tentar ABA 13, depois ABA 12, ABA 11...)
    # saiu de cena. A prioridade de sala agora não é mais uma ordem fixa —
    # é decidida sala a sala, na hora, com base no andar (mezanino/térreo)
    # e na preferência/restrição de cada assistido. Ver `_ordenar_por_sala`
    # e o novo bloco de alocação mais abaixo.
    #
    # `TODAS_AS_SALAS` é só o universo de salas conhecidas do sistema —
    # sem preferência nenhuma embutida. Se um dia existir sala nova, ela só
    # precisa entrar aqui (ou vir de config_pacientes.json, se preferir
    # manter configurável).
    TODAS_AS_SALAS = ["ABA 01", "ABA 02", "ABA 03", "ABA 04", "ABA 05", "ABA 06",
                       "ABA 07", "ABA 08", "ABA 09", "ABA 10", "ABA 11", "ABA 12", "ABA 13"]
    todas_as_salas_config = regras_gerais.get("todas_as_salas", TODAS_AS_SALAS)
    salas_ativas = [s for s in todas_as_salas_config if s not in bloqueadas]

    def _numero_sala(nome_sala):
        """Extrai o número da sala pra ordenação (ex: 'ABA 09' -> 9)."""
        try:
            return int(nome_sala.split()[1])
        except (IndexError, ValueError):
            return 0

    # Salas ativas separadas por andar, em ordem crescente de número.
    # A ordem entre salas do MESMO andar não importa pro resultado (Ken
    # confirmou) — ficou crescente só pra ser determinística e fácil de
    # depurar (rodar a mesma escala duas vezes dá o mesmo resultado).
    salas_mezanino_ativas = sorted([s for s in salas_ativas if s not in salas_terreo], key=_numero_sala)
    salas_terreo_ativas = sorted([s for s in salas_ativas if s in salas_terreo], key=_numero_sala)

    horarios = ["13:15", "14:00", "14:45", "15:30", "16:15", "17:00", "17:45"]
    mapa_final = {sala: {h: "" for h in horarios} for sala in salas_ativas + bloqueadas}

    pacientes_alocados_no_horario = {h: set() for h in horarios}
    nao_alocados = []

    if not isinstance(lista_pacientes, list):
        print(f"{DEBUG_TAG} ERRO: lista_pacientes não é uma lista (provavelmente veio uma mensagem de erro do processar_escala): {lista_pacientes}")
        return mapa_final, nao_alocados

    def _chave_prioridade_fila(item):
        """
        Ordem de prioridade da fila (confirmada com o Ken):
        1º sala fixa           - precisa da sala dela, sem excessão
        2º resistência a escada - necessidade física, não pode ficar sem sala
        3º não divide sala      - perfil exige sala exclusiva, também não
                                   pode ficar sem alocação
        4º prioridade clínica
        5º resto (o campo grupo_match só entra aqui pra decidir se DIVIDE
           bem com quem já está numa sala, não pra furar fila)
        """
        nome_bruto = item['nome'].strip().upper()
        nome = apelidos.get(nome_bruto, nome_bruto)
        info_paciente = pacientes_db.get(nome, {})
        return (
            not info_paciente.get('sala_fixa', ''),
            not info_paciente.get('resistencia_escada', False),
            info_paciente.get('divide_sala', True),
            not info_paciente.get('prioridade_clinica', False),
            not info_paciente.get('grupo_match'),
        )

    fila = sorted(lista_pacientes, key=_chave_prioridade_fila)

    for p in fila:
        nome_original = p['nome'].strip().upper()
        nome = apelidos.get(nome_original, nome_original)

        horario = p['horario'].replace("H", "").strip()
        info = pacientes_db.get(nome, {})
        if nome in pacientes_alocados_no_horario[horario]:
            continue

        sala_destinada = None

        # --- REGRA DO JORGE: assistidos dele vão sempre pra ABA 04, ---
        # --- ignorando bloqueio de sala e qualquer outra regra. ------
        # CORRIGIDO (Parte 2, item 3): usa `nome` (já corrigido pelo
        # apelidos) em vez de `nome_original` ao ESCREVER no mapa final.
        # Antes, o dicionário `apelidos` só corrigia a busca interna no
        # banco — o texto final que ia pro WhatsApp continuava mostrando
        # o erro de digitação (ex: "LUCCA CALVACANTI" em vez de "LUCCA
        # CAVALCANTI"), como confirmado num teste real em produção.
        if p.get("profissional") and _normalizar(p["profissional"]) == "JORGE":
            conteudo_atual_aba04 = mapa_final["ABA 04"][horario]
            if not conteudo_atual_aba04:
                mapa_final["ABA 04"][horario] = nome
            else:
                ja_tem = nome in conteudo_atual_aba04
                if not ja_tem:
                    qtd_atual = len(conteudo_atual_aba04.split(" / "))
                    mapa_final["ABA 04"][horario] = f"{conteudo_atual_aba04} / {nome}"
                    if qtd_atual + 1 > 2:
                        print(f"{DEBUG_TAG} AVISO: mais de 2 assistidos do Jorge na ABA 04 às {horario} — confira se está correto: {mapa_final['ABA 04'][horario]}")
            sala_destinada = "ABA 04"
            pacientes_alocados_no_horario[horario].add(nome)
            continue

        sf = info.get("sala_fixa")
        if sf and sf in mapa_final and sf not in bloqueadas:
            if not mapa_final[sf][horario] or nome in mapa_final[sf][horario]:
                sala_destinada = sf
                mapa_final[sf][horario] = nome

        if not sala_destinada:
            # "NO FLOW" (item 2): a ordem de tentativa agora é decidida por
            # andar, não mais por uma lista fixa de salas.
            # - Resistência a escada: só pode ir pro térreo (regra que já
            #   existia, mantida sem mudança nenhuma).
            # - Todo mundo mais — com `preferencia_mezanino` true OU false,
            #   Ken confirmou tratar os dois casos igual — tenta o mezanino
            #   primeiro, e só cai pro térreo se não tiver opção lá.
            if info.get("resistencia_escada"):
                ordem_andares = [salas_terreo_ativas]
            else:
                ordem_andares = [salas_mezanino_ativas, salas_terreo_ativas]

            # --- FASE 1: sala vazia, respeitando a ordem de andar acima ---
            for grupo_salas in ordem_andares:
                if sala_destinada:
                    break
                for sala in grupo_salas:
                    if not mapa_final[sala][horario]:
                        sala_destinada = sala
                        mapa_final[sala][horario] = nome
                        break

            # --- FASE 2: nenhuma sala vazia -> tenta dividir, priorizando
            # mezanino antes de térreo de novo (mesma ordem de andar) ---
            if not sala_destinada:
                for grupo_salas in ordem_andares:
                    if sala_destinada:
                        break
                    for sala in grupo_salas:
                        conteudo_atual = mapa_final[sala][horario]
                        if not conteudo_atual:
                            continue

                        dono_nome_bruto = conteudo_atual.split(" / ")[0].strip()
                        dono_nome = apelidos.get(dono_nome_bruto.upper(), dono_nome_bruto.upper())
                        info_dono = pacientes_db.get(dono_nome, {})

                        pode_entrar = not info_dono.get("sala_fixa") and info_dono.get("divide_sala", True)

                        # Máximo 2 por sala no automático (3 só na tela
                        # manual de alocar sem sala).
                        if pode_entrar and len(conteudo_atual.split(" / ")) < 2:
                            mesmo_grupo = (
                                info.get("grupo_match") is not None
                                and info.get("grupo_match") == info_dono.get("grupo_match")
                            )
                            if mesmo_grupo:
                                mapa_final[sala][horario] = f"{conteudo_atual} / {nome}"
                                sala_destinada = sala
                                break

        # --- FASE 3 (fallback): prioridade clínica pode dividir mesmo sem
        # bater grupo_match, como último recurso — respeitando a mesma
        # ordem de andar e o máximo de 2 por sala no automático. ---
        if not sala_destinada and info.get("prioridade_clinica"):
            ordem_andares = [salas_terreo_ativas] if info.get("resistencia_escada") else [salas_mezanino_ativas, salas_terreo_ativas]
            for grupo_salas in ordem_andares:
                if sala_destinada:
                    break
                for sala in grupo_salas:
                    conteudo_atual = mapa_final[sala][horario]
                    if conteudo_atual:
                        dono_nome_bruto = conteudo_atual.split(" / ")[0].strip()
                        dono_nome = apelidos.get(dono_nome_bruto.upper(), dono_nome_bruto.upper())
                        info_dono = pacientes_db.get(dono_nome, {})

                        if not info_dono.get("sala_fixa") and info_dono.get("divide_sala", True):
                            if len(conteudo_atual.split(" / ")) < 2:
                                mapa_final[sala][horario] = f"{conteudo_atual} / {nome}"
                                sala_destinada = sala
                                break

        if sala_destinada:
            if " / " not in mapa_final[sala_destinada][horario]:
                mapa_final[sala_destinada][horario] = nome
            pacientes_alocados_no_horario[horario].add(nome)
        else:
            nao_alocados.append(f"{horario} - {nome}")

    print(f"{DEBUG_TAG} distribuir_salas_ia() concluída. {len(nao_alocados)} pacientes sem sala.")
    return mapa_final, nao_alocados


# --- FORMATAÇÃO PARA TEXTO FINAL ---

# --- ESCRITA DIRETA NA PLANILHA DE VACÂNCIA (Ponto 3) ---

# Mesma ordem de horário usada em distribuir_salas_ia() / formatar_mapa_para_texto(),
# sem o sufixo "H" (o "H" só existe no rótulo visual da coluna A da Vacância).
HORARIOS_VACANCIA = ["13:15", "14:00", "14:45", "15:30", "16:15", "17:00", "17:45"]


def _coluna_sala_vacancia(indice_sala):
    """Sala 1 = coluna B, Sala 2 = C, ..., Sala 13 = N."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(1 + indice_sala)


def escrever_vacancia(url_vacancia, nome_aba, mapa_final):
    """
    Escreve o mapa de alocação (mesmo formato de distribuir_salas_ia():
    "ABA 01".."ABA 13" -> horário -> nome) na aba do dia correspondente da
    planilha de Vacância, via Sheets API (values_update, uma chamada só).

    Mexe SÓ nas 13 colunas de sala normal (B até N, "Sala 1".."Sala 13").
    As colunas de Musicoterapia com Tatames / Caixa de Areia / Mercado da
    Inclusão ficam de fora por enquanto — dependem do Ponto 4 (ainda não
    implementado), que vai definir como esses destinos entram no mapa.

    Unidade de trabalho: só a aba `nome_aba` é tocada, nunca as outras —
    1:1 com o dia selecionado na leitura do Direcionamento (Ponto 1).

    Sem log e sem backup antes de escrever — decisão explícita do Ken pra
    manter essa primeira versão simples (ver handoff #9, Ponto 3).
    """
    print(f"{DEBUG_TAG} escrever_vacancia() iniciada. aba={nome_aba}")

    planilha = conectar_google_sheets(url_vacancia)
    try:
        aba = planilha.worksheet(nome_aba)
    except Exception:
        raise ValueError(f"Aba '{nome_aba}' não encontrada na planilha de Vacância.")

    valores = []
    for horario in HORARIOS_VACANCIA:
        linha = []
        for n in range(1, 14):
            sala_key = f"ABA {n:02d}"
            nome = mapa_final.get(sala_key, {}).get(horario, "")
            linha.append(nome)
        valores.append(linha)

    col_inicio = _coluna_sala_vacancia(1)   # B
    col_fim = _coluna_sala_vacancia(13)     # N
    linha_inicio = 3   # primeira linha de horário (13:15H)
    linha_fim = 2 + len(HORARIOS_VACANCIA)  # última linha de horário (17:45H)
    range_destino = f"{nome_aba}!{col_inicio}{linha_inicio}:{col_fim}{linha_fim}"

    aba.spreadsheet.values_update(
        range_destino,
        params={"valueInputOption": "RAW"},
        body={"values": valores},
    )

    total_celulas = len(valores) * len(valores[0]) if valores else 0
    print(f"{DEBUG_TAG} Vacância atualizada: '{range_destino}' ({total_celulas} células).")
    return total_celulas


def formatar_mapa_para_texto(mapa_final, nao_alocados):
    """Transforma o dicionário de salas no texto final formatado para o usuário."""
    print(f"{DEBUG_TAG} formatar_mapa_para_texto() chamada.")
    texto = "📌*Deem prioridade aos menores estar no MEZANINO. As ABAS de baixo são prioritárias dos assistidos com resistências as escadas!*\n\n"
    texto += "*VACÂNCIA DIÁRIA*✨💚\n\n"

    horarios = ["13:15", "14:00", "14:45", "15:30", "16:15", "17:00", "17:45"]
    salas_ordenadas = sorted(mapa_final.keys(), key=lambda x: int(x.split()[1]))

    for sala in salas_ordenadas:
        eh_normalmente_bloqueada = "04" in sala or "06" in sala
        tem_ocupante = any(mapa_final[sala].get(h) for h in horarios)

        if eh_normalmente_bloqueada and not tem_ocupante:
            icon = "❌"
        elif eh_normalmente_bloqueada and tem_ocupante:
            # Bloqueada pra uso geral, mas tem gente lá (ex: exceção do Jorge
            # na ABA 04) — precisa aparecer, senão a equipe não vê quem tá lá.
            icon = "🔒"
        else:
            icon = "✨"

        texto += f"*{sala}* {icon}\n"

        if icon == "❌":
            texto += "\n"
            continue

        for h in horarios:
            paciente = mapa_final[sala].get(h, "")
            texto += f"{h} {paciente}\n"
        texto += "\n"

    texto += "---\n⚠️ *ALERTA: PACIENTES SEM SALA (LOTADO)*\n"

    if not nao_alocados:
        texto += "✅ Todos os pacientes foram alocados com sucesso."
    else:
        for item in nao_alocados:
            if isinstance(item, dict):
                h_p = item.get('horario', '??:??')
                n_p = item.get('nome', 'Desconhecido')
                texto += f"• {h_p} - {n_p}\n"
            else:
                texto += f"• Horário não definido - {item}\n"

    return texto
